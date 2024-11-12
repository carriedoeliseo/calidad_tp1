# -*- coding: utf-8 -*-

import openpyxl
import pandas as pd
from inline_sql import sql
from openpyxl.styles import PatternFill

#%% ===========================================================================

df = pd.read_excel('./tiburones_color.xlsx')

#%% ===========================================================================

'''
• Unprovoked Incidents = #eddab5
• Provoked Incidents = #ffcc99 
• Attacks on Boats = #cfeecc
• Air / Sea Disasters = #ffffc5
• Questionable Incidents = #99ccff
'''

no_match_color = df[ (df['HexColor'] != '#eddab5') & 
                     (df['HexColor'] != '#ffcc99') &
                     (df['HexColor'] != '#cfeecc') &
                     (df['HexColor'] != '#fef2cb') &
                     (df['HexColor'] != '#99ccff') &
                     (df['HexColor'] != '#fbe4d5') &
                     (df['HexColor'] != '#ffffc5') ]


#%% ===========================================================================

wb = openpyxl.load_workbook('./tiburon2.xlsx')
sheet = wb.active

#%% ===========================================================================

df = df[['Type', 'Country', 'State', 'Location', 'Injury']]
def obtener_color_fondo(celda):
    # Verifica si la celda tiene un color de fondo
    color = celda.fill.start_color
    if color.type == "rgb":  # Si el color está en formato RGB
        return f"#{color.rgb[2:]}"  # Elimina los primeros dos caracteres (transparencia)
    elif color.type == "theme":  # Si usa un color de tema
        return f"Tema-{color.theme}"
    return None  # Si no hay color o el tipo es desconocido

colores_fondo = []

for fila in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    fila_colores_fondo = [obtener_color_fondo(celda) for celda in fila]
    colores_fondo.append(fila_colores_fondo)

colores_fondo = pd.DataFrame()

#%% ===========================================================================

data0 = pd.read_excel('./Tiburones-ultimo.xlsx')
data0 = data0[['Id', 'Date', 'Fecha transformada', 'Color', 'Incidente', 'Year',
       'Type', 'Coincide tipo', 'Country', 'Columna reconciliada', 'State',
       'Location', 'Activity', 'Name', 'Sex', 'Age', 'Injury', 'COUNT',
       'FATAL', 'Column', 'Time', 'Species', 'Source', 'pdf', 'href formula',
       'href', 'Case Number', 'Case Number 2', 'original order']]

data0.columns = ['Id', 'Date', 'Fecha', 'Color', 'Incidente', 'Year',
       'Type', 'Coincide tipo', 'Country', 'Pais', 'State',
       'Location', 'Activity', 'Name', 'Sex', 'Age', 'Injury', 'COUNT',
       'FATAL', 'Column', 'Time', 'Species', 'Source', 'pdf', 'href formula',
       'href', 'Case Number', 'Case Number 2', 'original order']

data = data0.dropna(subset=['Fecha', 'Incidente', 'Pais', 'Location' ])

data = data.set_index('Id', drop=False)

duplicados = sql^ """
                 SELECT *
                 FROM data AS f1
                 WHERE EXISTS (SELECT 1
                               FROM data AS f2
                               WHERE f1.Fecha = f2.Fecha AND
                                     f1.Pais = f2.Pais AND
                                     f1.Incidente = f2.Incidente AND
                                     f1.Location = f2.Location AND
                                     f1.Id != f2.Id)
                 ORDER BY f1.Fecha, f1.Pais
                 """
                 
duplicados2 = sql^ """
                 SELECT Fecha, Incidente, Pais, Location, COUNT(*)
                 FROM data AS f1
                 WHERE EXISTS (SELECT 1
                               FROM data AS f2
                               WHERE f1.Fecha = f2.Fecha AND
                                     f1.Pais = f2.Pais AND
                                     f1.Incidente = f2.Incidente AND
                                     f1.Location = f2.Location AND
                                     f1.Id != f2.Id)
                 GROUP BY f1.Fecha, f1.Incidente, f1.Pais, f1.Location
                 ORDER BY f1.Fecha, f1.Pais
                 """

# Borra los FALSOS que tienen duplicado VERDADERO y TODOS los VERDADEROS duplicados
#a_borrar1 = sql^ """
 #                SELECT *
  #               FROM data AS f1
   #              WHERE EXISTS (SELECT 1
    #                           FROM data AS f2
     #                          WHERE f1."Fecha Transformada" = f2."Fecha Transformada" AND
      #                               f1.Country = f2.Country AND
       #                              f1.Type = f2.Type AND
        #                             f1.Location = f2.Location AND
         #                            f1.Id != f2.Id AND
          #                           f2.FATAL = 'VERDADERO' LIMIT 1)
           #      ORDER BY f1."Fecha Transformada", f1.Country
            #     """

# Borra los FALSOS que tienen duplicado VERDADERO
a_borrar2 = sql^ """
                 SELECT *
                 FROM data AS f1
                 WHERE EXISTS (SELECT 1
                               FROM data AS f2
                               WHERE f1.Fecha = f2.Fecha AND
                                     f1.Pais = f2.Pais AND
                                     f1.Incidente = f2.Incidente AND
                                     f1.Location = f2.Location AND
                                     f1.Id != f2.Id AND
                                     f2.FATAL = 'VERDADERO') AND f1.FATAL != 'VERDADERO'
                 ORDER BY f1.Fecha, f1.Pais
                 """
                 
a_borrar2 = a_borrar2.set_index('Id', drop=False)

todos_los_borrados = a_borrar2[['Id']]

data_borrados1 = data.drop(a_borrar2.index)
data_borrados2 = data_borrados1[['Id', 'Date', 'Fecha', 'Incidente', 'Pais', 'Location', 'FATAL']]

duplicados_borrar = sql^ """
                         SELECT DISTINCT Id
                         FROM (SELECT *, ROW_NUMBER() OVER(PARTITION BY Fecha,
                                                                        Pais,
                                                                        Incidente,
                                                                        Location) rn
                               FROM data_borrados2) alias
                         WHERE rn != 1
                         """

# SON 132 DE 6917
todos_los_borrados = pd.concat([duplicados_borrar[['Id']], todos_los_borrados])

#%% ===========================================================================

data0 = data0.set_index('Id', drop=False)
data0 = data0.drop(todos_los_borrados['Id'])
data0 = data0.drop(columns=['COUNT'])

data0.to_excel('./tiburones_sin_duplicados.xlsx')

#%% ===========================================================================

consulta_provisional = sql^ """
                            SELECT DISTINCT *
                            FROM data0 AS d
                            WHERE d.Column = 'Y'
                            """
                            
consulta_provisional2 = sql^ """
                            SELECT DISTINCT FATAL, COUNT(*)
                            FROM consulta_provisional
                            GROUP BY FATAL
                            """
                            
                            